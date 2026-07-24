#!/usr/bin/env python3
"""
Create Appendix Tables A1–A3 from FGSS sensitivity-analysis CSV files.

The workbook contains:
    Table A1 — FGSS 1 (PBU)
    Table A2 — FGSS 2 (Pump)
    Table A3 — FGSS 3 (Compressor)

Table structure
---------------
Rows:
    All numeric performance indicators contained in the sensitivity-result CSV.

Columns:
    Four varied parameters, each divided into Low / Base / High levels:
        - Insulation, λ/λ₀
        - Initial tank filling
        - Ambient temperature
        - Fuel-demand scaling

The workbook is formatted for convenient transfer into a Systems (MDPI)
manuscript:
    - editable Excel cells
    - Arial 8 pt body text
    - appendix numbering
    - multi-level headers
    - horizontal rules only
    - wrapped headings
    - landscape print setup
    - repeated header rows when printed
    - no decorative charts or unnecessary colors

Required package
----------------
Install openpyxl once in the active Python environment:

    python -m pip install openpyxl

Typical project structure
-------------------------
project/
├── results/
│   └── sensitivity/
│       ├── PBU_dual_sensitivity_results.csv
│       ├── Pump_dual_sensitivity_results.csv
│       └── Compressor_dual_sensitivity_results.csv
└── scripts/
    └── analysis/
        └── create_appendix_sensitivity_tables_openpyxl.py

The default path logic assumes the script is in scripts/analysis and therefore
uses the parent of the parent of SCRIPT_DIR as PROJECT_ROOT.
"""

from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# =============================================================================
# USER SETTINGS
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

DATA_DIR = PROJECT_ROOT / "results" / "sensitivity"
OUTPUT_DIR = PROJECT_ROOT / "results" / "sensitivity"
OUTPUT_XLSX = OUTPUT_DIR / "Appendix_A_sensitivity_tables.xlsx"

# Set one or more of these to explicit Path objects when automatic detection
# should be disabled, for example:
#
# PBU_CSV = DATA_DIR / "PBU_dual_sensitivity_results(4).csv"
#
PBU_CSV: Path | None = None
PUMP_CSV: Path | None = None
COMPRESSOR_CSV: Path | None = None

FONT_NAME = "Arial"
BODY_FONT_SIZE = 8
HEADER_FONT_SIZE = 8
CAPTION_FONT_SIZE = 9

# Set True to hide performance indicators that contain only zeros or blanks
# for a given FGSS configuration. The default keeps all numeric indicators.
HIDE_ALL_ZERO_ROWS = False


# =============================================================================
# TABLE DEFINITIONS
# =============================================================================

PARAMETER_ORDER = [
    "tank_ins_scale",
    "initial_fill",
    "T_amb",
    "fuel_flow_scale",
]

PARAMETER_LABELS = {
    "tank_ins_scale": "Insulation, λ/λ₀",
    "initial_fill": "Initial tank filling",
    "T_amb": "Ambient temperature",
    "fuel_flow_scale": "Fuel-demand scaling",
}

LEVEL_LABELS = {
    "low": "Low",
    "base": "Base",
    "high": "High",
}

REQUIRED_COLUMNS = {
    "parameter",
    "level",
    "value",
    "termination_reason",
}

EXCLUDED_COLUMNS = {
    "parameter",
    "level",
    "value",
    "termination_reason",
    "heel_reached",
    "mawp_exceeded",
    "results_truncated_at_mawp",
}

# A thin horizontal separator is inserted above these indicators.
SECTION_START_COLUMNS = {
    "final_fill_pct",
    "total_bog_removed_kg",
    "requested_fuel_kg",
    "pump_energy_kWh",
    "startup_time_days",
    "bog_balance_error_kg",
}


@dataclass(frozen=True)
class IndicatorSpec:
    label: str
    unit: str
    factor: float = 1.0
    number_format: str = "0.000"


INDICATOR_SPECS: dict[str, IndicatorSpec] = {
    "simulation_duration_days": IndicatorSpec(
        "Simulation duration", "d"
    ),
    "voyage_simulation_duration_days": IndicatorSpec(
        "Voyage simulation duration", "d"
    ),
    "time_to_min_heel_days": IndicatorSpec(
        "Time to minimum heel", "d"
    ),
    "elapsed_time_to_min_heel_days": IndicatorSpec(
        "Elapsed time to minimum heel", "d"
    ),
    "voyage_time_to_min_heel_days": IndicatorSpec(
        "Voyage time to minimum heel", "d"
    ),
    "time_to_mawp_days": IndicatorSpec(
        "Time to MAWP", "d"
    ),
    "final_fill_pct": IndicatorSpec(
        "Final tank filling", "%"
    ),
    "max_pressure_bar": IndicatorSpec(
        "Maximum tank pressure", "bar"
    ),
    "pressure_at_termination_bar": IndicatorSpec(
        "Tank pressure at termination", "bar"
    ),
    "pressure_margin_to_mawp_bar": IndicatorSpec(
        "Pressure margin to MAWP", "bar"
    ),
    "total_bog_removed_kg": IndicatorSpec(
        "Total BOG removed", "t", factor=0.001
    ),
    "bog_used_as_fuel_kg": IndicatorSpec(
        "BOG used as fuel", "t", factor=0.001
    ),
    "bog_excess_kg": IndicatorSpec(
        "Excess BOG requiring management", "t", factor=0.001
    ),
    "bog_removed_lng_kg": IndicatorSpec(
        "BOG removed during LNG operation", "t", factor=0.001
    ),
    "bog_removed_conventional_kg": IndicatorSpec(
        "BOG removed during conventional-fuel operation", "t", factor=0.001
    ),
    "conventional_operation_days": IndicatorSpec(
        "Conventional-fuel operation", "d"
    ),
    "requested_fuel_kg": IndicatorSpec(
        "Requested fuel", "t", factor=0.001
    ),
    "fuel_supplied_kg": IndicatorSpec(
        "Fuel supplied", "t", factor=0.001
    ),
    "lng_supply_shortfall_kg": IndicatorSpec(
        "LNG-supply shortfall", "t", factor=0.001
    ),
    "unserved_fuel_kg": IndicatorSpec(
        "Unserved fuel", "t", factor=0.001
    ),
    "pump_energy_kWh": IndicatorSpec(
        "Pump mechanical energy", "kWh", number_format="0.0"
    ),
    "compressor_energy_kWh": IndicatorSpec(
        "Compressor mechanical energy", "kWh", number_format="0.0"
    ),
    "mechanical_energy_kWh": IndicatorSpec(
        "Total mechanical energy", "kWh", number_format="0.0"
    ),
    "specific_mechanical_energy_kWh_per_t": IndicatorSpec(
        "Specific mechanical energy", "kWh t⁻¹"
    ),
    "pbu_thermal_energy_kWh": IndicatorSpec(
        "PBU thermal energy", "kWh", number_format="0.0"
    ),
    "startup_pbu_thermal_energy_kWh": IndicatorSpec(
        "Startup PBU thermal energy", "kWh", number_format="0.0"
    ),
    "operational_pbu_thermal_energy_kWh": IndicatorSpec(
        "Operational PBU thermal energy", "kWh", number_format="0.0"
    ),
    "evaporator_evaporation_energy_kWh": IndicatorSpec(
        "Evaporator evaporation energy", "kWh", number_format="0.0"
    ),
    "evaporator_internal_superheat_energy_kWh": IndicatorSpec(
        "Evaporator internal-superheating energy", "kWh", number_format="0.0"
    ),
    "evaporator_thermal_energy_kWh": IndicatorSpec(
        "Total evaporator thermal energy", "kWh", number_format="0.0"
    ),
    "superheater_thermal_energy_kWh": IndicatorSpec(
        "Superheater thermal energy", "kWh", number_format="0.0"
    ),
    "total_process_thermal_energy_kWh": IndicatorSpec(
        "Total process thermal energy", "kWh", number_format="0.0"
    ),
    "startup_time_days": IndicatorSpec(
        "Startup time", "d"
    ),
    "pressure_recovery_time_days": IndicatorSpec(
        "Pressure-recovery time", "d"
    ),
    "pbu_pressurization_time_days": IndicatorSpec(
        "PBU pressurization time", "d"
    ),
    "bog_balance_error_kg": IndicatorSpec(
        "BOG mass-balance error", "kg", number_format="0.000E+00"
    ),
}


@dataclass(frozen=True)
class TableSpec:
    sheet_name: str
    table_number: str
    system_name: str
    file_stem: str
    explicit_path: Path | None


TABLE_SPECS = (
    TableSpec(
        sheet_name="Table A1",
        table_number="A1",
        system_name="FGSS 1 (PBU)",
        file_stem="PBU_dual_sensitivity_results",
        explicit_path=PBU_CSV,
    ),
    TableSpec(
        sheet_name="Table A2",
        table_number="A2",
        system_name="FGSS 2 (Pump)",
        file_stem="Pump_dual_sensitivity_results",
        explicit_path=PUMP_CSV,
    ),
    TableSpec(
        sheet_name="Table A3",
        table_number="A3",
        system_name="FGSS 3 (Compressor)",
        file_stem="Compressor_dual_sensitivity_results",
        explicit_path=COMPRESSOR_CSV,
    ),
)


# =============================================================================
# CSV READING AND TRANSFORMATION
# =============================================================================

def parse_float(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if text == "":
        return None

    try:
        number = float(text)
    except ValueError:
        return None

    return number if math.isfinite(number) else None


def csv_has_required_columns(path: Path) -> bool:
    try:
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            header = next(reader)
    except (OSError, StopIteration, UnicodeError):
        return False

    return REQUIRED_COLUMNS.issubset(set(header))


def file_version(path: Path, stem: str) -> int:
    if path.name == f"{stem}.csv":
        return 0

    match = re.fullmatch(
        rf"{re.escape(stem)}\((\d+)\)\.csv",
        path.name,
    )
    return int(match.group(1)) if match else -1


def resolve_csv_file(
    data_dir: Path,
    stem: str,
    explicit_path: Path | None,
) -> Path:
    if explicit_path is not None:
        path = explicit_path.expanduser().resolve()

        if not path.is_file():
            raise FileNotFoundError(f"CSV file not found: {path}")

        if not csv_has_required_columns(path):
            raise ValueError(
                f"{path.name} does not contain the required sensitivity columns."
            )

        return path

    candidates = [
        path
        for path in data_dir.glob(f"{stem}*.csv")
        if csv_has_required_columns(path)
    ]

    if not candidates:
        raise FileNotFoundError(
            f"No valid sensitivity-result CSV matching '{stem}*.csv' "
            f"was found in:\n  {data_dir}"
        )

    # Prefer the most recently modified valid file. The numeric suffix is used
    # as a secondary criterion when modification times are identical.
    return max(
        candidates,
        key=lambda path: (
            path.stat().st_mtime,
            file_version(path, stem),
        ),
    )


def read_csv_records(
    path: Path,
) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)

        if reader.fieldnames is None:
            raise ValueError(f"No header was found in {path.name}.")

        rows = list(reader)

    return list(reader.fieldnames), rows


def is_numeric_indicator(
    column: str,
    rows: list[dict[str, str]],
) -> bool:
    if column in EXCLUDED_COLUMNS:
        return False

    return any(
        parse_float(row.get(column)) is not None
        for row in rows
    )


def humanize_column_name(column: str) -> str:
    replacements = {
        "pbu": "PBU",
        "bog": "BOG",
        "lng": "LNG",
        "mawp": "MAWP",
        "kwh": "kWh",
        "pct": "%",
    }

    formatted: list[str] = []

    for word in column.split("_"):
        formatted.append(
            replacements.get(word.lower(), word)
        )

    label = " ".join(formatted)
    return label[:1].upper() + label[1:]


def infer_indicator_spec(column: str) -> IndicatorSpec:
    label = humanize_column_name(column)

    if column.endswith("_days"):
        return IndicatorSpec(label.removesuffix(" days"), "d")

    if column.endswith("_pct"):
        return IndicatorSpec(label.removesuffix(" %"), "%")

    if column.endswith("_bar"):
        return IndicatorSpec(label.removesuffix(" bar"), "bar")

    if column.endswith("_kWh_per_t"):
        return IndicatorSpec(
            label.removesuffix(" kWh per t"),
            "kWh t⁻¹",
        )

    if column.endswith("_kWh"):
        return IndicatorSpec(
            label.removesuffix(" kWh"),
            "kWh",
            number_format="0.0",
        )

    if column.endswith("_kg"):
        return IndicatorSpec(
            label.removesuffix(" kg"),
            "kg",
        )

    if column.endswith("_K"):
        return IndicatorSpec(
            label.removesuffix(" K"),
            "K",
        )

    if column.endswith("_s"):
        return IndicatorSpec(
            label.removesuffix(" s"),
            "s",
        )

    return IndicatorSpec(label, "–")


def ordered_indicator_columns(
    fieldnames: list[str],
    rows: list[dict[str, str]],
) -> list[str]:
    numeric_columns = [
        column
        for column in fieldnames
        if is_numeric_indicator(column, rows)
    ]

    known = [
        column
        for column in INDICATOR_SPECS
        if column in numeric_columns
    ]

    extras = [
        column
        for column in numeric_columns
        if column not in INDICATOR_SPECS
    ]

    return known + extras


def get_global_baseline_row(
    rows: list[dict[str, str]],
) -> dict[str, str]:
    matches = [
        row
        for row in rows
        if row.get("parameter") == "baseline"
    ]

    if len(matches) != 1:
        raise ValueError(
            "Each CSV must contain exactly one row with parameter='baseline'."
        )

    return matches[0]


def get_parameter_level_row(
    rows: list[dict[str, str]],
    parameter: str,
    level: str,
    baseline_row: dict[str, str],
) -> dict[str, str]:
    matches = [
        row
        for row in rows
        if row.get("parameter") == parameter
        and row.get("level") == level
    ]

    if len(matches) == 1:
        return matches[0]

    if level == "base":
        return baseline_row

    raise ValueError(
        f"Expected one row for parameter={parameter!r}, "
        f"level={level!r}; found {len(matches)}."
    )


def format_parameter_value(
    parameter: str,
    raw_value: object,
) -> str:
    value = parse_float(raw_value)

    if value is None:
        return ""

    if parameter == "tank_ins_scale":
        return f"{value:.2f}"

    if parameter == "initial_fill":
        return f"{100.0 * value:.0f}%"

    if parameter == "T_amb":
        return f"{value:.2f} K"

    if parameter == "fuel_flow_scale":
        return f"{value:.2f}"

    return f"{value:g}"


def transform_value(
    row: dict[str, str],
    indicator: str,
    spec: IndicatorSpec,
) -> float | None:
    value = parse_float(row.get(indicator))

    if value is None:
        return None

    return value * spec.factor


def values_are_all_zero_or_blank(
    values: Iterable[float | None],
) -> bool:
    finite_values = [
        value
        for value in values
        if value is not None
    ]

    if not finite_values:
        return True

    return all(math.isclose(value, 0.0, abs_tol=1e-12) for value in finite_values)


# =============================================================================
# EXCEL FORMATTING
# =============================================================================

THIN_BLACK = Side(
    style="thin",
    color="000000",
)

HAIR_GRAY = Side(
    style="hair",
    color="808080",
)

NO_SIDE = Side(
    style=None,
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2",
)


def apply_top_border(
    worksheet,
    row: int,
    first_column: int,
    last_column: int,
    side: Side,
) -> None:
    for column in range(first_column, last_column + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=side,
            bottom=cell.border.bottom,
        )


def apply_bottom_border(
    worksheet,
    row: int,
    first_column: int,
    last_column: int,
    side: Side,
) -> None:
    for column in range(first_column, last_column + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=side,
        )


def style_worksheet(
    worksheet,
    last_column: int,
    last_data_row: int,
    note_row: int,
) -> None:
    last_column_letter = get_column_letter(last_column)

    # Global font and vertical alignment.
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=note_row,
        min_col=1,
        max_col=last_column,
    ):
        for cell in row:
            cell.font = Font(
                name=FONT_NAME,
                size=BODY_FONT_SIZE,
                color="000000",
            )
            cell.alignment = Alignment(
                vertical="center",
            )

    # Caption.
    caption_cell = worksheet["A1"]
    caption_cell.font = Font(
        name=FONT_NAME,
        size=CAPTION_FONT_SIZE,
        bold=True,
        color="000000",
    )
    caption_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[1].height = 26

    # Two header rows.
    for row_number in (3, 4):
        for cell in worksheet[row_number]:
            if cell.column > last_column:
                continue

            cell.font = Font(
                name=FONT_NAME,
                size=HEADER_FONT_SIZE,
                bold=True,
                color="000000",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.fill = HEADER_FILL

    worksheet.row_dimensions[3].height = 24
    worksheet.row_dimensions[4].height = 34

    # Body alignment.
    for row_number in range(5, last_data_row + 1):
        worksheet.cell(row=row_number, column=1).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        worksheet.cell(row=row_number, column=2).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for column in range(3, last_column + 1):
            worksheet.cell(
                row=row_number,
                column=column,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.row_dimensions[row_number].height = 18

    # Note.
    note_cell = worksheet.cell(row=note_row, column=1)
    note_cell.font = Font(
        name=FONT_NAME,
        size=BODY_FONT_SIZE,
        italic=True,
        color="000000",
    )
    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[note_row].height = 42

    # Widths.
    worksheet.column_dimensions["A"].width = 39
    worksheet.column_dimensions["B"].width = 11

    for column in range(3, last_column + 1):
        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = 11.5

    # Horizontal rules only.
    apply_top_border(
        worksheet,
        row=3,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )
    apply_bottom_border(
        worksheet,
        row=4,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )
    apply_bottom_border(
        worksheet,
        row=last_data_row,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )

    # Freeze and printing.
    worksheet.freeze_panes = "C5"
    worksheet.sheet_view.showGridLines = False

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.45,
        bottom=0.45,
        header=0.15,
        footer=0.15,
    )

    worksheet.print_title_rows = "3:4"
    worksheet.print_area = f"A1:{last_column_letter}{note_row}"
    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False

    # Center the table horizontally on the printed page.
    worksheet.print_options.horizontalCentered = True


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def build_table_sheet(
    workbook: Workbook,
    table_spec: TableSpec,
    csv_path: Path,
) -> None:
    fieldnames, rows = read_csv_records(csv_path)
    baseline_row = get_global_baseline_row(rows)
    indicators = ordered_indicator_columns(
        fieldnames,
        rows,
    )

    if not indicators:
        raise ValueError(
            f"No numeric performance indicators were found in {csv_path.name}."
        )

    worksheet = workbook.create_sheet(
        title=table_spec.sheet_name,
    )

    number_of_columns = 2 + 3 * len(PARAMETER_ORDER)
    last_column = number_of_columns
    last_column_letter = get_column_letter(last_column)

    caption = (
        f"Table {table_spec.table_number}. Complete one-at-a-time "
        f"sensitivity-analysis results for {table_spec.system_name}."
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    worksheet["A1"] = caption

    # First two header cells span both header rows.
    worksheet.merge_cells("A3:A4")
    worksheet.merge_cells("B3:B4")
    worksheet["A3"] = "Performance indicator"
    worksheet["B3"] = "Unit"

    parameter_rows: dict[
        tuple[str, str],
        dict[str, str],
    ] = {}

    has_mawp_limited_case = False
    current_column = 3

    for parameter in PARAMETER_ORDER:
        start_column = current_column
        end_column = current_column + 2

        worksheet.merge_cells(
            start_row=3,
            start_column=start_column,
            end_row=3,
            end_column=end_column,
        )
        worksheet.cell(
            row=3,
            column=start_column,
            value=PARAMETER_LABELS[parameter],
        )

        for offset, level in enumerate(
            ("low", "base", "high")
        ):
            row = get_parameter_level_row(
                rows=rows,
                parameter=parameter,
                level=level,
                baseline_row=baseline_row,
            )
            parameter_rows[(parameter, level)] = row

            marker = ""

            if (
                str(row.get("termination_reason", ""))
                .strip()
                .upper()
                == "MAWP"
            ):
                marker = "†"
                has_mawp_limited_case = True

            parameter_value = format_parameter_value(
                parameter,
                row.get("value"),
            )

            if parameter_value:
                header = (
                    f"{LEVEL_LABELS[level]}{marker}\n"
                    f"({parameter_value})"
                )
            else:
                header = f"{LEVEL_LABELS[level]}{marker}"

            worksheet.cell(
                row=4,
                column=start_column + offset,
                value=header,
            )

        current_column += 3

    first_data_row = 5
    output_rows: list[
        tuple[str, IndicatorSpec, list[float | None]]
    ] = []

    for indicator in indicators:
        indicator_spec = INDICATOR_SPECS.get(
            indicator,
            infer_indicator_spec(indicator),
        )

        numeric_values: list[float | None] = []

        for parameter in PARAMETER_ORDER:
            for level in ("low", "base", "high"):
                source_row = parameter_rows[
                    (parameter, level)
                ]
                numeric_values.append(
                    transform_value(
                        source_row,
                        indicator,
                        indicator_spec,
                    )
                )

        if (
            HIDE_ALL_ZERO_ROWS
            and values_are_all_zero_or_blank(numeric_values)
        ):
            continue

        output_rows.append(
            (
                indicator,
                indicator_spec,
                numeric_values,
            )
        )

    for row_offset, (
        indicator,
        indicator_spec,
        numeric_values,
    ) in enumerate(output_rows):
        excel_row = first_data_row + row_offset

        worksheet.cell(
            row=excel_row,
            column=1,
            value=indicator_spec.label,
        )
        worksheet.cell(
            row=excel_row,
            column=2,
            value=indicator_spec.unit,
        )

        for value_offset, value in enumerate(
            numeric_values,
            start=3,
        ):
            cell = worksheet.cell(
                row=excel_row,
                column=value_offset,
                value=value,
            )
            cell.number_format = indicator_spec.number_format

        if (
            indicator in SECTION_START_COLUMNS
            and excel_row > first_data_row
        ):
            apply_top_border(
                worksheet,
                row=excel_row,
                first_column=1,
                last_column=last_column,
                side=HAIR_GRAY,
            )

    last_data_row = (
        first_data_row + len(output_rows) - 1
    )

    note_row = last_data_row + 2
    note_parts = [
        "Note: Each parameter was varied individually while all other "
        "parameters were kept at their baseline values.",
        "Blank cells indicate that the corresponding indicator was not "
        "available or that the relevant termination event was not reached.",
        f"Source file: {csv_path.name}.",
    ]

    if has_mawp_limited_case:
        note_parts.append(
            "† Simulation terminated at the maximum allowable working "
            "pressure (MAWP) before the minimum heel was reached."
        )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=last_column,
    )
    worksheet.cell(
        row=note_row,
        column=1,
        value=" ".join(note_parts),
    )

    style_worksheet(
        worksheet=worksheet,
        last_column=last_column,
        last_data_row=last_data_row,
        note_row=note_row,
    )


def create_workbook() -> Workbook:
    workbook = Workbook()

    # Remove the blank default worksheet.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table_spec in TABLE_SPECS:
        csv_path = resolve_csv_file(
            data_dir=DATA_DIR,
            stem=table_spec.file_stem,
            explicit_path=table_spec.explicit_path,
        )

        print(
            f"{table_spec.sheet_name}: "
            f"{table_spec.system_name} <- {csv_path.name}"
        )

        build_table_sheet(
            workbook=workbook,
            table_spec=table_spec,
            csv_path=csv_path,
        )

    return workbook


def main() -> None:
    if not DATA_DIR.is_dir():
        raise FileNotFoundError(
            "Sensitivity-results directory was not found:\n"
            f"  {DATA_DIR}\n\n"
            "Edit DATA_DIR in the USER SETTINGS section."
        )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = create_workbook()
    workbook.save(OUTPUT_XLSX)

    print(
        f"\nWorkbook saved to:\n"
        f"  {OUTPUT_XLSX.resolve()}"
    )


if __name__ == "__main__":
    main()
