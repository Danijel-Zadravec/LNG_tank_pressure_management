#!/usr/bin/env python3
"""
Create compact Appendix Tables A1–A3 from FGSS sensitivity-analysis CSV files.

Compact table structure
-----------------------
Each worksheet contains only five columns:

    Performance indicator | Insulation | Initial filling | Ambient temperature | Fuel demand

Each parameter cell reports the three one-at-a-time results as:

    low | base | high

The actual low, base, and high parameter settings are included in the column
header. Units are included directly in the performance-indicator name, so no
separate unit column is used.

Required package
----------------
Install openpyxl in the active virtual environment:

    python -m pip install openpyxl

Expected project structure
--------------------------
project/
├── results/
│   └── sensitivity/
│       ├── PBU_dual_sensitivity_results.csv
│       ├── Pump_dual_sensitivity_results.csv
│       └── Compressor_dual_sensitivity_results.csv
└── scripts/
    └── analysis/
        └── create_appendix_sensitivity_tables_compact.py

The output workbook is written to:

    results/sensitivity/Appendix_A_sensitivity_tables_compact.xlsx
"""

from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# =============================================================================
# USER SETTINGS
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

DATA_DIR = PROJECT_ROOT / "results" / "sensitivity"
OUTPUT_DIR = PROJECT_ROOT / "results" / "sensitivity"
OUTPUT_XLSX = OUTPUT_DIR / "Appendix_A_sensitivity_tables_compact.xlsx"

# Explicit paths may be supplied when automatic detection should be disabled.
PBU_CSV: Path | None = None
PUMP_CSV: Path | None = None
COMPRESSOR_CSV: Path | None = None

FONT_NAME = "Arial"
BODY_FONT_SIZE = 8
HEADER_FONT_SIZE = 8
CAPTION_FONT_SIZE = 9

# Keep all numeric indicators by default. Set True to omit rows that contain
# only zeros and blanks for the relevant FGSS configuration.
HIDE_ALL_ZERO_ROWS = False

VALUE_SEPARATOR = " | "
MISSING_VALUE = "—"


# =============================================================================
# TABLE DEFINITIONS
# =============================================================================

PARAMETER_ORDER = [
    "tank_ins_scale",
    "initial_fill",
    "T_amb",
    "fuel_flow_scale",
]

PARAMETER_LABELS = {
    "tank_ins_scale": "Insulation, λ/λ₀",
    "initial_fill": "Initial filling",
    "T_amb": "Ambient temperature",
    "fuel_flow_scale": "Fuel-demand scaling",
}

LEVEL_ORDER = ("low", "base", "high")

REQUIRED_COLUMNS = {
    "parameter",
    "level",
    "value",
    "termination_reason",
}

EXCLUDED_COLUMNS = {
    "parameter",
    "level",
    "value",
    "termination_reason",
    "heel_reached",
    "mawp_exceeded",
    "results_truncated_at_mawp",
}

# Insert subtle horizontal separators above the first item of each group.
SECTION_START_COLUMNS = {
    "final_fill_pct",
    "total_bog_removed_kg",
    "requested_fuel_kg",
    "pump_energy_kWh",
    "startup_time_days",
    "bog_balance_error_kg",
}


@dataclass(frozen=True)
class IndicatorSpec:
    label: str
    unit: str
    factor: float = 1.0
    decimals: int = 3
    scientific: bool = False

    @property
    def display_name(self) -> str:
        if self.unit == "–":
            return self.label
        return f"{self.label} ({self.unit})"


# Short manuscript-oriented indicator names.
INDICATOR_SPECS: dict[str, IndicatorSpec] = {
    "simulation_duration_days": IndicatorSpec(
        "Simulation duration", "d"
    ),
    "voyage_simulation_duration_days": IndicatorSpec(
        "Voyage duration", "d"
    ),
    "time_to_min_heel_days": IndicatorSpec(
        "Time to min. heel", "d"
    ),
    "elapsed_time_to_min_heel_days": IndicatorSpec(
        "Elapsed time to min. heel", "d"
    ),
    "voyage_time_to_min_heel_days": IndicatorSpec(
        "Voyage time to min. heel", "d"
    ),
    "time_to_mawp_days": IndicatorSpec(
        "Time to MAWP", "d"
    ),
    "final_fill_pct": IndicatorSpec(
        "Final filling", "%", decimals=2
    ),
    "max_pressure_bar": IndicatorSpec(
        "Max. tank pressure", "bar"
    ),
    "pressure_at_termination_bar": IndicatorSpec(
        "Pressure at termination", "bar"
    ),
    "pressure_margin_to_mawp_bar": IndicatorSpec(
        "Margin to MAWP", "bar"
    ),
    "total_bog_removed_kg": IndicatorSpec(
        "Total BOG removed", "t", factor=0.001
    ),
    "bog_used_as_fuel_kg": IndicatorSpec(
        "BOG used as fuel", "t", factor=0.001
    ),
    "bog_excess_kg": IndicatorSpec(
        "Excess BOG", "t", factor=0.001
    ),
    "bog_removed_lng_kg": IndicatorSpec(
        "BOG removed—LNG mode", "t", factor=0.001
    ),
    "bog_removed_conventional_kg": IndicatorSpec(
        "BOG removed—conventional mode", "t", factor=0.001
    ),
    "conventional_operation_days": IndicatorSpec(
        "Conventional-fuel operation", "d"
    ),
    "requested_fuel_kg": IndicatorSpec(
        "Fuel requested", "t", factor=0.001
    ),
    "fuel_supplied_kg": IndicatorSpec(
        "Fuel supplied", "t", factor=0.001
    ),
    "lng_supply_shortfall_kg": IndicatorSpec(
        "LNG-supply shortfall", "t", factor=0.001
    ),
    "unserved_fuel_kg": IndicatorSpec(
        "Unserved fuel", "t", factor=0.001
    ),
    "pump_energy_kWh": IndicatorSpec(
        "Pump energy", "kWh", decimals=1
    ),
    "compressor_energy_kWh": IndicatorSpec(
        "Compressor energy", "kWh", decimals=1
    ),
    "mechanical_energy_kWh": IndicatorSpec(
        "Total mechanical energy", "kWh", decimals=1
    ),
    "specific_mechanical_energy_kWh_per_t": IndicatorSpec(
        "Specific mechanical energy", "kWh t⁻¹"
    ),
    "pbu_thermal_energy_kWh": IndicatorSpec(
        "PBU thermal energy", "kWh", decimals=1
    ),
    "startup_pbu_thermal_energy_kWh": IndicatorSpec(
        "Startup PBU energy", "kWh", decimals=1
    ),
    "operational_pbu_thermal_energy_kWh": IndicatorSpec(
        "Operating PBU energy", "kWh", decimals=1
    ),
    "evaporator_evaporation_energy_kWh": IndicatorSpec(
        "Evaporation energy", "kWh", decimals=1
    ),
    "evaporator_internal_superheat_energy_kWh": IndicatorSpec(
        "Internal superheating energy", "kWh", decimals=1
    ),
    "evaporator_thermal_energy_kWh": IndicatorSpec(
        "Evaporator thermal energy", "kWh", decimals=1
    ),
    "superheater_thermal_energy_kWh": IndicatorSpec(
        "Superheater energy", "kWh", decimals=1
    ),
    "total_process_thermal_energy_kWh": IndicatorSpec(
        "Total process heat", "kWh", decimals=1
    ),
    "startup_time_days": IndicatorSpec(
        "Startup time", "d"
    ),
    "pressure_recovery_time_days": IndicatorSpec(
        "Pressure-recovery time", "d"
    ),
    "pbu_pressurization_time_days": IndicatorSpec(
        "PBU pressurization time", "d"
    ),
    "bog_balance_error_kg": IndicatorSpec(
        "BOG balance error", "kg", scientific=True
    ),
}


@dataclass(frozen=True)
class TableSpec:
    sheet_name: str
    table_number: str
    system_name: str
    file_stem: str
    explicit_path: Path | None


TABLE_SPECS = (
    TableSpec(
        sheet_name="Table A1",
        table_number="A1",
        system_name="FGSS 1 (PBU)",
        file_stem="PBU_dual_sensitivity_results",
        explicit_path=PBU_CSV,
    ),
    TableSpec(
        sheet_name="Table A2",
        table_number="A2",
        system_name="FGSS 2 (Pump)",
        file_stem="Pump_dual_sensitivity_results",
        explicit_path=PUMP_CSV,
    ),
    TableSpec(
        sheet_name="Table A3",
        table_number="A3",
        system_name="FGSS 3 (Compressor)",
        file_stem="Compressor_dual_sensitivity_results",
        explicit_path=COMPRESSOR_CSV,
    ),
)


# =============================================================================
# CSV READING AND TRANSFORMATION
# =============================================================================

def parse_float(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        number = float(text)
    except ValueError:
        return None

    return number if math.isfinite(number) else None


def read_header(path: Path) -> list[str] | None:
    try:
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            return next(csv.reader(handle))
    except (OSError, StopIteration, UnicodeError):
        return None


def csv_has_required_columns(path: Path) -> bool:
    header = read_header(path)
    return header is not None and REQUIRED_COLUMNS.issubset(set(header))


def suffix_version(path: Path, stem: str) -> int:
    match = re.fullmatch(
        rf"{re.escape(stem)}(?:\((\d+)\))?\.csv",
        path.name,
    )
    if not match:
        return -1
    return int(match.group(1) or 0)


def resolve_csv_file(
    data_dir: Path,
    stem: str,
    explicit_path: Path | None,
) -> Path:
    if explicit_path is not None:
        path = explicit_path.expanduser().resolve()

        if not path.is_file():
            raise FileNotFoundError(f"CSV file not found: {path}")

        if not csv_has_required_columns(path):
            raise ValueError(
                f"{path.name} is not a valid sensitivity-result CSV."
            )

        return path

    candidates: list[Path] = []

    for path in data_dir.glob(f"{stem}*.csv"):
        if csv_has_required_columns(path):
            candidates.append(path)

    if not candidates:
        raise FileNotFoundError(
            f"No valid sensitivity-result CSV matching '{stem}*.csv' "
            f"was found in:\n  {data_dir}"
        )

    # Prefer the richest valid export; use suffix and modification time to
    # resolve ties. This avoids selecting an older reduced-format CSV.
    return max(
        candidates,
        key=lambda path: (
            len(read_header(path) or []),
            suffix_version(path, stem),
            path.stat().st_mtime,
        ),
    )


def read_csv_records(
    path: Path,
) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)

        if reader.fieldnames is None:
            raise ValueError(f"No header was found in {path.name}.")

        rows = list(reader)

    return list(reader.fieldnames), rows


def is_numeric_indicator(
    column: str,
    rows: list[dict[str, str]],
) -> bool:
    if column in EXCLUDED_COLUMNS:
        return False

    return any(
        parse_float(row.get(column)) is not None
        for row in rows
    )


def humanize_column_name(column: str) -> str:
    replacements = {
        "pbu": "PBU",
        "bog": "BOG",
        "lng": "LNG",
        "mawp": "MAWP",
        "kwh": "kWh",
        "pct": "%",
    }

    words = [
        replacements.get(word.lower(), word)
        for word in column.split("_")
    ]

    label = " ".join(words)
    return label[:1].upper() + label[1:]


def infer_indicator_spec(column: str) -> IndicatorSpec:
    label = humanize_column_name(column)

    if column.endswith("_days"):
        return IndicatorSpec(label.removesuffix(" days"), "d")

    if column.endswith("_pct"):
        return IndicatorSpec(label.removesuffix(" %"), "%", decimals=2)

    if column.endswith("_bar"):
        return IndicatorSpec(label.removesuffix(" bar"), "bar")

    if column.endswith("_kWh_per_t"):
        return IndicatorSpec(
            label.removesuffix(" kWh per t"),
            "kWh t⁻¹",
        )

    if column.endswith("_kWh"):
        return IndicatorSpec(
            label.removesuffix(" kWh"),
            "kWh",
            decimals=1,
        )

    if column.endswith("_kg"):
        return IndicatorSpec(
            label.removesuffix(" kg"),
            "kg",
        )

    if column.endswith("_K"):
        return IndicatorSpec(
            label.removesuffix(" K"),
            "K",
        )

    if column.endswith("_s"):
        return IndicatorSpec(
            label.removesuffix(" s"),
            "s",
        )

    return IndicatorSpec(label, "–")


def ordered_indicator_columns(
    fieldnames: list[str],
    rows: list[dict[str, str]],
) -> list[str]:
    numeric_columns = [
        column
        for column in fieldnames
        if is_numeric_indicator(column, rows)
    ]

    known = [
        column
        for column in INDICATOR_SPECS
        if column in numeric_columns
    ]

    extras = [
        column
        for column in numeric_columns
        if column not in INDICATOR_SPECS
    ]

    return known + extras


def get_global_baseline_row(
    rows: list[dict[str, str]],
) -> dict[str, str]:
    matches = [
        row
        for row in rows
        if row.get("parameter") == "baseline"
    ]

    if len(matches) != 1:
        raise ValueError(
            "Each CSV must contain exactly one row with parameter='baseline'."
        )

    return matches[0]


def get_parameter_level_row(
    rows: list[dict[str, str]],
    parameter: str,
    level: str,
    baseline_row: dict[str, str],
) -> dict[str, str]:
    matches = [
        row
        for row in rows
        if row.get("parameter") == parameter
        and row.get("level") == level
    ]

    if len(matches) == 1:
        return matches[0]

    if level == "base":
        return baseline_row

    raise ValueError(
        f"Expected one row for parameter={parameter!r}, "
        f"level={level!r}; found {len(matches)}."
    )


def format_parameter_value(
    parameter: str,
    raw_value: object,
) -> str:
    value = parse_float(raw_value)

    if value is None:
        return MISSING_VALUE

    if parameter == "tank_ins_scale":
        return f"{value:.2f}"

    if parameter == "initial_fill":
        return f"{100.0 * value:.0f}%"

    if parameter == "T_amb":
        return f"{value:.2f} K"

    if parameter == "fuel_flow_scale":
        return f"{value:.2f}"

    return f"{value:g}"


def format_result_value(
    raw_value: object,
    spec: IndicatorSpec,
) -> str:
    value = parse_float(raw_value)

    if value is None:
        return MISSING_VALUE

    value *= spec.factor

    if spec.scientific:
        return f"{value:.3E}"

    if math.isclose(value, 0.0, abs_tol=10 ** (-(spec.decimals + 1))):
        return "0"

    return f"{value:.{spec.decimals}f}"


def values_are_all_zero_or_blank(
    raw_values: Iterable[object],
    spec: IndicatorSpec,
) -> bool:
    converted: list[float] = []

    for raw_value in raw_values:
        value = parse_float(raw_value)
        if value is not None:
            converted.append(value * spec.factor)

    if not converted:
        return True

    return all(
        math.isclose(value, 0.0, abs_tol=1e-12)
        for value in converted
    )


# =============================================================================
# EXCEL FORMATTING
# =============================================================================

THIN_BLACK = Side(style="thin", color="000000")
HAIR_GRAY = Side(style="hair", color="808080")

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2",
)


def apply_top_border(
    worksheet,
    row: int,
    first_column: int,
    last_column: int,
    side: Side,
) -> None:
    for column in range(first_column, last_column + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=side,
            bottom=cell.border.bottom,
        )


def apply_bottom_border(
    worksheet,
    row: int,
    first_column: int,
    last_column: int,
    side: Side,
) -> None:
    for column in range(first_column, last_column + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=side,
        )


def style_worksheet(
    worksheet,
    last_data_row: int,
    note_row: int,
) -> None:
    last_column = 5

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=note_row,
        min_col=1,
        max_col=last_column,
    ):
        for cell in row:
            cell.font = Font(
                name=FONT_NAME,
                size=BODY_FONT_SIZE,
                color="000000",
            )
            cell.alignment = Alignment(vertical="center")

    # Caption.
    worksheet["A1"].font = Font(
        name=FONT_NAME,
        size=CAPTION_FONT_SIZE,
        bold=True,
        color="000000",
    )
    worksheet["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[1].height = 29

    # Compact single header row.
    for cell in worksheet[3]:
        if cell.column > last_column:
            continue

        cell.font = Font(
            name=FONT_NAME,
            size=HEADER_FONT_SIZE,
            bold=True,
            color="000000",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.fill = HEADER_FILL

    worksheet.row_dimensions[3].height = 42

    # Body.
    for row_number in range(4, last_data_row + 1):
        worksheet.cell(row=row_number, column=1).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        for column in range(2, last_column + 1):
            worksheet.cell(
                row=row_number,
                column=column,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[row_number].height = 18

    # Note.
    worksheet.cell(row=note_row, column=1).font = Font(
        name=FONT_NAME,
        size=BODY_FONT_SIZE,
        italic=True,
        color="000000",
    )
    worksheet.cell(row=note_row, column=1).alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[note_row].height = 46

    # Compact widths.
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 22
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 23
    worksheet.column_dimensions["E"].width = 22

    # Horizontal rules only.
    apply_top_border(
        worksheet,
        row=3,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )
    apply_bottom_border(
        worksheet,
        row=3,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )
    apply_bottom_border(
        worksheet,
        row=last_data_row,
        first_column=1,
        last_column=last_column,
        side=THIN_BLACK,
    )

    worksheet.freeze_panes = "B4"
    worksheet.sheet_view.showGridLines = False

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.45,
        bottom=0.45,
        header=0.15,
        footer=0.15,
    )

    worksheet.print_title_rows = "3:3"
    worksheet.print_area = f"A1:E{note_row}"
    worksheet.print_options.horizontalCentered = True


# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def build_table_sheet(
    workbook: Workbook,
    table_spec: TableSpec,
    csv_path: Path,
) -> None:
    fieldnames, rows = read_csv_records(csv_path)
    baseline_row = get_global_baseline_row(rows)
    indicators = ordered_indicator_columns(fieldnames, rows)

    if not indicators:
        raise ValueError(
            f"No numeric performance indicators were found in {csv_path.name}."
        )

    worksheet = workbook.create_sheet(title=table_spec.sheet_name)

    parameter_rows: dict[
        tuple[str, str],
        dict[str, str],
    ] = {}

    has_mawp_limited_case = False

    for parameter in PARAMETER_ORDER:
        for level in LEVEL_ORDER:
            row = get_parameter_level_row(
                rows=rows,
                parameter=parameter,
                level=level,
                baseline_row=baseline_row,
            )
            parameter_rows[(parameter, level)] = row

            if (
                str(row.get("termination_reason", ""))
                .strip()
                .upper()
                == "MAWP"
            ):
                has_mawp_limited_case = True

    caption = (
        f"Table {table_spec.table_number}. Complete one-at-a-time "
        f"sensitivity-analysis results for {table_spec.system_name}. "
        f"Values in each parameter column are reported as "
        f"low{VALUE_SEPARATOR}base{VALUE_SEPARATOR}high."
    )

    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = caption

    worksheet["A3"] = "Performance indicator"

    for column_offset, parameter in enumerate(
        PARAMETER_ORDER,
        start=2,
    ):
        setting_values = []

        for level in LEVEL_ORDER:
            row = parameter_rows[(parameter, level)]
            setting = format_parameter_value(
                parameter,
                row.get("value"),
            )

            if (
                str(row.get("termination_reason", ""))
                .strip()
                .upper()
                == "MAWP"
            ):
                setting += "†"

            setting_values.append(setting)

        worksheet.cell(
            row=3,
            column=column_offset,
            value=(
                f"{PARAMETER_LABELS[parameter]}\n"
                f"({VALUE_SEPARATOR.join(setting_values)})"
            ),
        )

    first_data_row = 4
    output_rows: list[
        tuple[str, IndicatorSpec, list[str], list[object]]
    ] = []

    for indicator in indicators:
        indicator_spec = INDICATOR_SPECS.get(
            indicator,
            infer_indicator_spec(indicator),
        )

        raw_values_for_zero_test: list[object] = []
        compact_cells: list[str] = []

        for parameter in PARAMETER_ORDER:
            level_values: list[str] = []

            for level in LEVEL_ORDER:
                source_row = parameter_rows[(parameter, level)]
                raw_value = source_row.get(indicator)
                raw_values_for_zero_test.append(raw_value)

                formatted = format_result_value(
                    raw_value,
                    indicator_spec,
                )

                if (
                    str(source_row.get("termination_reason", ""))
                    .strip()
                    .upper()
                    == "MAWP"
                ):
                    formatted += "†"

                level_values.append(formatted)

            compact_cells.append(
                VALUE_SEPARATOR.join(level_values)
            )

        if (
            HIDE_ALL_ZERO_ROWS
            and values_are_all_zero_or_blank(
                raw_values_for_zero_test,
                indicator_spec,
            )
        ):
            continue

        output_rows.append(
            (
                indicator,
                indicator_spec,
                compact_cells,
                raw_values_for_zero_test,
            )
        )

    for row_offset, (
        indicator,
        indicator_spec,
        compact_cells,
        _,
    ) in enumerate(output_rows):
        excel_row = first_data_row + row_offset

        worksheet.cell(
            row=excel_row,
            column=1,
            value=indicator_spec.display_name,
        )

        for column_offset, value in enumerate(
            compact_cells,
            start=2,
        ):
            worksheet.cell(
                row=excel_row,
                column=column_offset,
                value=value,
            )

        if (
            indicator in SECTION_START_COLUMNS
            and excel_row > first_data_row
        ):
            apply_top_border(
                worksheet,
                row=excel_row,
                first_column=1,
                last_column=5,
                side=HAIR_GRAY,
            )

    last_data_row = first_data_row + len(output_rows) - 1
    note_row = last_data_row + 2

    notes = [
        f"Note: Values are reported as low{VALUE_SEPARATOR}base"
        f"{VALUE_SEPARATOR}high for the parameter identified in each column.",
        "Each parameter was varied individually while all other parameters "
        "were kept at their baseline values.",
        "A dash indicates that the indicator was unavailable or the relevant "
        "termination event was not reached.",
    ]

    if has_mawp_limited_case:
        notes.append(
            "† The corresponding simulation reached the maximum allowable "
            "working pressure (MAWP) before the minimum heel."
        )

    worksheet.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=5,
    )
    worksheet.cell(
        row=note_row,
        column=1,
        value=" ".join(notes),
    )

    style_worksheet(
        worksheet=worksheet,
        last_data_row=last_data_row,
        note_row=note_row,
    )


def create_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for table_spec in TABLE_SPECS:
        csv_path = resolve_csv_file(
            data_dir=DATA_DIR,
            stem=table_spec.file_stem,
            explicit_path=table_spec.explicit_path,
        )

        print(
            f"{table_spec.sheet_name}: "
            f"{table_spec.system_name} <- {csv_path.name}"
        )

        build_table_sheet(
            workbook=workbook,
            table_spec=table_spec,
            csv_path=csv_path,
        )

    return workbook


def main() -> None:
    if not DATA_DIR.is_dir():
        raise FileNotFoundError(
            "Sensitivity-results directory was not found:\n"
            f"  {DATA_DIR}\n\n"
            "Edit DATA_DIR in the USER SETTINGS section."
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = create_workbook()
    workbook.save(OUTPUT_XLSX)

    print(
        f"\nWorkbook saved to:\n"
        f"  {OUTPUT_XLSX.resolve()}"
    )


if __name__ == "__main__":
    main()
